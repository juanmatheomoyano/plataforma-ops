import io
import logging
import uuid
from datetime import datetime, timezone

import httpx
import openpyxl

logger = logging.getLogger(__name__)
from fastapi import HTTPException, UploadFile, status
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.core.security import decrypt, encrypt

from . import baproar_client
from .models import EstadoKeys, Seller
from .schemas import ImportError, SellerCreate, SellerImportResult, SellerImportUpdateResult, SellerUpdate

# Excel column → model field mapping (lower-stripped keys for fuzzy match)
_COL_MAP = {
    "id seller_ecommerce": "id_ecommerce",
    "nombre de fantasía": "seller_name",
    "nombre de fantasia": "seller_name",
    "seller id": "seller_id",
    "app key": "app_key",
    "app token": "app_token",
    "usuario que la creó": "creado_por",
    "usuario que la creo": "creado_por",
    "fecha de creación": "fecha_creacion",
    "fecha de creacion": "fecha_creacion",
    "estado keys": "estado_keys",
    "integración": "integracion",
    "integracion": "integracion",
    "vendiendo": "vendiendo",
    "analista": "analista",
    "notas / observaciones": "notas",
    "notas": "notas",
}


def _map_col(header: str) -> str | None:
    return _COL_MAP.get(header.strip().lower())


def _parse_bool(val: str | bool | None) -> bool:
    if isinstance(val, bool):
        return val
    if val is None:
        return False
    return str(val).strip().lower() in {"sí", "si", "yes", "true", "1", "s"}


def _parse_estado(val: str | None) -> EstadoKeys:
    if not val:
        return EstadoKeys.activo
    v = str(val).strip().lower()
    for e in EstadoKeys:
        if e.value == v:
            return e
    return EstadoKeys.activo


async def get_all_sellers(
    db: AsyncSession,
    skip: int = 0,
    limit: int = 200,
) -> list[Seller]:
    result = await db.execute(
        select(Seller).order_by(Seller.created_at.desc()).offset(skip).limit(limit)
    )
    return list(result.scalars().all())


async def get_seller_by_id(seller_id: uuid.UUID, db: AsyncSession) -> Seller:
    result = await db.execute(select(Seller).where(Seller.id == seller_id))
    seller = result.scalar_one_or_none()
    if not seller:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Seller not found")
    return seller


async def create_seller(data: SellerCreate, db: AsyncSession) -> Seller:
    seller = Seller(
        id_ecommerce=data.id_ecommerce,
        seller_name=data.seller_name,
        seller_id=data.seller_id,
        app_key_enc=encrypt(data.app_key),
        app_token_enc=encrypt(data.app_token),
        creado_por=data.creado_por,
        fecha_creacion=data.fecha_creacion,
        estado_keys=data.estado_keys,
        integracion=data.integracion,
        integracion_spec=data.integracion_spec,
        vendiendo=data.vendiendo,
        analista=data.analista,
        notas=data.notas,
    )
    db.add(seller)
    try:
        await db.commit()
        await db.refresh(seller)
    except IntegrityError:
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="id_ecommerce or seller_id already exists",
        )
    return seller


async def update_seller(
    seller_id: uuid.UUID, data: SellerUpdate, db: AsyncSession
) -> Seller:
    seller = await get_seller_by_id(seller_id, db)

    # exclude_unset=True: solo los campos presentes en el body; permite setear null explícito
    update_data = data.model_dump(exclude_unset=True)

    if "app_key" in update_data:
        seller.app_key_enc = encrypt(update_data.pop("app_key"))
    else:
        update_data.pop("app_key", None)

    if "app_token" in update_data:
        seller.app_token_enc = encrypt(update_data.pop("app_token"))
    else:
        update_data.pop("app_token", None)

    for field, value in update_data.items():
        setattr(seller, field, value)

    seller.updated_at = datetime.now(timezone.utc)

    try:
        await db.commit()
        await db.refresh(seller)
    except IntegrityError:
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="id_ecommerce already in use",
        )
    return seller


async def deactivate_seller(seller_id: uuid.UUID, db: AsyncSession) -> Seller:
    seller = await get_seller_by_id(seller_id, db)
    seller.is_active = False
    seller.updated_at = datetime.now(timezone.utc)
    await db.commit()
    await db.refresh(seller)
    return seller


def get_decrypted_credentials(seller: Seller) -> tuple[str, str]:
    """Internal use only — never expose via API."""
    return decrypt(seller.app_key_enc), decrypt(seller.app_token_enc)


async def test_connection(seller_id: uuid.UUID, db: AsyncSession) -> dict:
    seller = await get_seller_by_id(seller_id, db)
    if not seller.app_key_enc or not seller.app_token_enc:
        return {"ok": False, "error": "Credenciales no configuradas"}

    try:
        app_key, app_token = get_decrypted_credentials(seller)
        url = (
            f"https://{seller.seller_id}.vtexcommercestable.com.br"
            "/api/catalog_system/pub/products/search"
        )
        async with httpx.AsyncClient(timeout=10.0) as client:
            resp = await client.get(
                url,
                params={"_from": "0", "_to": "1"},
                headers={
                    "X-VTEX-API-AppKey": app_key,
                    "X-VTEX-API-AppToken": app_token,
                },
            )
        if resp.status_code < 400:
            return {"ok": True}
        return {"ok": False, "error": f"HTTP {resp.status_code}"}
    except httpx.TimeoutException:
        return {"ok": False, "error": "Timeout al conectar con VTEX"}
    except Exception as e:
        return {"ok": False, "error": str(e)}


async def import_sellers_from_file(
    file: UploadFile, db: AsyncSession
) -> SellerImportResult:
    content = await file.read()
    import io
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return SellerImportResult(total=0, exitosos=0, errores=0, detalle_errores=[])

    # Map header → field name
    raw_headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    headers = [_map_col(h) for h in raw_headers]

    exitosos = 0
    errors: list[ImportError] = []

    for row_idx, raw_row in enumerate(rows[1:], start=2):
        row_dict: dict = {}
        for col_idx, field in enumerate(headers):
            if field and col_idx < len(raw_row):
                row_dict[field] = raw_row[col_idx]

        sid = str(row_dict.get("seller_id") or "").strip()

        try:
            data = SellerCreate(
                id_ecommerce=str(row_dict.get("id_ecommerce") or "").strip(),
                seller_name=str(row_dict.get("seller_name") or "").strip(),
                seller_id=sid,
                app_key=str(row_dict.get("app_key") or "").strip(),
                app_token=str(row_dict.get("app_token") or "").strip(),
                creado_por=str(row_dict.get("creado_por") or "").strip() or None,
                fecha_creacion=str(row_dict.get("fecha_creacion") or "").strip() or None,
                estado_keys=_parse_estado(row_dict.get("estado_keys")),
                integracion=str(row_dict.get("integracion") or "").strip() or None,
                vendiendo=_parse_bool(row_dict.get("vendiendo")),
                analista=str(row_dict.get("analista") or "").strip() or None,
                notas=str(row_dict.get("notas") or "").strip() or None,
            )
            await create_seller(data, db)
            exitosos += 1
        except HTTPException as e:
            errors.append(ImportError(fila=row_idx, seller_id=sid or None, motivo=e.detail))
        except Exception as e:
            errors.append(ImportError(fila=row_idx, seller_id=sid or None, motivo=str(e)))

    total = len(rows) - 1
    return SellerImportResult(
        total=total,
        exitosos=exitosos,
        errores=len(errors),
        detalle_errores=errors,
    )


_EXPORT_COLS = [
    "id_ecommerce", "seller_name", "seller_id", "app_key", "app_token",
    "analista", "estado_keys", "integracion", "integracion_spec",
    "vendiendo", "creado_por", "fecha_creacion", "notas", "is_active",
]

_EXPORT_HEADERS = [
    "id_ecommerce", "seller_name", "seller_id", "App Key", "App Token",
    "analista", "estado_keys", "integracion", "integracion_spec",
    "vendiendo", "creado_por", "fecha_creacion", "notas", "is_active",
]


def _map_col_update(header: str) -> str | None:
    h = header.strip().lower()
    if h in _EXPORT_COLS:
        return h
    return _COL_MAP.get(h)


def _parse_active(val) -> bool:
    if isinstance(val, bool):
        return val
    return str(val).strip().lower() in {"activo", "active", "true", "1", "sí", "si", "s"}


async def sync_marketplace_sellers(db: AsyncSession) -> dict:
    """Llama a BaproAR, actualiza marketplace_activo en los sellers que coincidan por seller_id."""
    if not settings.BAPROAR_APP_KEY or not settings.BAPROAR_APP_TOKEN:
        logger.warning("sync_marketplace_sellers: BAPROAR_APP_KEY/TOKEN no configuradas — sync omitido")
        return {"synced": 0, "total_marketplace": 0, "error": "Credenciales BaproAR no configuradas"}

    try:
        marketplace_sellers = await baproar_client.list_sellers(
            settings.BAPROAR_APP_KEY, settings.BAPROAR_APP_TOKEN
        )
    except Exception as e:
        logger.error("sync_marketplace_sellers: error al llamar BaproAR: %s", e)
        return {"synced": 0, "total_marketplace": 0, "error": str(e)}

    # Construir mapa account → {isActive, baproar_id}
    # El campo "account" es el nombre de cuenta VTEX del seller (coincide con seller_id en nuestra BD)
    # El campo "id" es el identificador interno de BaproAR (a veces taxCode, a veces el account)
    account_map: dict[str, dict] = {}
    for ms in marketplace_sellers:
        account = ms.get("account")
        is_active = ms.get("isActive")
        baproar_id = ms.get("id")
        if isinstance(account, str) and account and isinstance(is_active, bool):
            account_map[account] = {"is_active": is_active, "baproar_id": str(baproar_id) if baproar_id else account}

    if not account_map:
        return {"synced": 0, "total_marketplace": len(marketplace_sellers), "error": "Sin sellers válidos en la respuesta"}

    now = datetime.now(timezone.utc)
    result = await db.execute(select(Seller).where(Seller.seller_id.in_(account_map.keys())))
    sellers_to_update = list(result.scalars().all())

    for seller in sellers_to_update:
        entry = account_map[seller.seller_id]
        seller.marketplace_activo = entry["is_active"]
        seller.marketplace_seller_id = entry["baproar_id"]
        seller.marketplace_sync_at = now
        seller.updated_at = now

    await db.commit()
    logger.info("sync_marketplace_sellers: %d sellers actualizados de %d en marketplace", len(sellers_to_update), len(account_map))
    return {"synced": len(sellers_to_update), "total_marketplace": len(status_map)}


async def toggle_marketplace_seller(seller_id: uuid.UUID, db: AsyncSession) -> Seller:
    """Activa/desactiva un seller en BaproAR y luego actualiza la BD solo si la API responde OK."""
    if not settings.BAPROAR_APP_KEY or not settings.BAPROAR_APP_TOKEN:
        raise HTTPException(status_code=503, detail="Credenciales BaproAR no configuradas")

    seller = await get_seller_by_id(seller_id, db)

    if seller.marketplace_activo is None:
        raise HTTPException(status_code=400, detail="Este seller no está registrado en el marketplace BaproAR")

    new_state = not seller.marketplace_activo

    baproar_id = seller.marketplace_seller_id or seller.seller_id
    try:
        await baproar_client.toggle_seller(
            baproar_id, new_state,
            settings.BAPROAR_APP_KEY, settings.BAPROAR_APP_TOKEN,
        )
    except httpx.HTTPStatusError as e:
        raise HTTPException(status_code=502, detail=f"BaproAR rechazó la operación: HTTP {e.response.status_code}")
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Error al contactar BaproAR: {e}")

    # Solo actualizamos la BD después de confirmar que BaproAR respondió OK
    seller.marketplace_activo = new_state
    seller.marketplace_sync_at = datetime.now(timezone.utc)
    seller.updated_at = datetime.now(timezone.utc)
    await db.commit()
    await db.refresh(seller)
    return seller


async def export_sellers_xlsx(db: AsyncSession) -> bytes:
    sellers = await get_all_sellers(db, limit=10000)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sellers"
    ws.append(_EXPORT_HEADERS)
    for s in sellers:
        try:
            app_key, app_token = get_decrypted_credentials(s) if s.app_key_enc else ("", "")
        except Exception:
            app_key, app_token = "", ""
        ws.append([
            s.id_ecommerce,
            s.seller_name,
            s.seller_id,
            app_key,
            app_token,
            s.analista,
            s.estado_keys.value if s.estado_keys else "",
            s.integracion,
            s.integracion_spec,
            "Sí" if s.vendiendo else "No",
            s.creado_por,
            s.fecha_creacion,
            s.notas,
            "activo" if s.is_active else "inactivo",
        ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def import_update_sellers(
    file: UploadFile, db: AsyncSession
) -> SellerImportUpdateResult:
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return SellerImportUpdateResult(total=0, actualizados=0, creados=0, errores=0, detalle_errores=[])

    raw_headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    col_idx_map = {i: _map_col_update(h) for i, h in enumerate(raw_headers) if _map_col_update(h)}

    actualizados = creados = 0
    errors: list[ImportError] = []

    for row_idx, raw_row in enumerate(rows[1:], start=2):
        row_dict: dict = {col_idx_map[i]: raw_row[i] for i in col_idx_map if i < len(raw_row)}
        sid = str(row_dict.get("seller_id") or "").strip()
        if not sid:
            continue

        try:
            result = await db.execute(select(Seller).where(Seller.seller_id == sid))
            existing = result.scalar_one_or_none()

            if existing:
                for field in ("id_ecommerce", "seller_name", "analista", "integracion",
                              "integracion_spec", "creado_por", "fecha_creacion", "notas"):
                    val = row_dict.get(field)
                    if val is not None:
                        setattr(existing, field, str(val).strip() or None if field not in ("id_ecommerce", "seller_name") else str(val).strip())
                if "estado_keys" in row_dict:
                    existing.estado_keys = _parse_estado(row_dict["estado_keys"])
                if "vendiendo" in row_dict:
                    existing.vendiendo = _parse_bool(row_dict["vendiendo"])
                if "is_active" in row_dict:
                    existing.is_active = _parse_active(row_dict["is_active"])
                ak = str(row_dict.get("app_key") or "").strip()
                at = str(row_dict.get("app_token") or "").strip()
                if ak:
                    existing.app_key_enc = encrypt(ak)
                if at:
                    existing.app_token_enc = encrypt(at)
                existing.updated_at = datetime.now(timezone.utc)
                await db.commit()
                await db.refresh(existing)
                actualizados += 1
            else:
                id_ecommerce = str(row_dict.get("id_ecommerce") or "").strip()
                seller_name = str(row_dict.get("seller_name") or "").strip()
                if not id_ecommerce or not seller_name:
                    errors.append(ImportError(fila=row_idx, seller_id=sid, motivo="id_ecommerce y seller_name son requeridos para crear"))
                    continue
                ak = str(row_dict.get("app_key") or "").strip()
                at = str(row_dict.get("app_token") or "").strip()
                new_seller = Seller(
                    id_ecommerce=id_ecommerce,
                    seller_name=seller_name,
                    seller_id=sid,
                    app_key_enc=encrypt(ak) if ak else None,
                    app_token_enc=encrypt(at) if at else None,
                    analista=str(row_dict.get("analista") or "").strip() or None,
                    estado_keys=_parse_estado(row_dict.get("estado_keys")),
                    integracion=str(row_dict.get("integracion") or "").strip() or None,
                    integracion_spec=str(row_dict.get("integracion_spec") or "").strip() or None,
                    vendiendo=_parse_bool(row_dict.get("vendiendo")),
                    creado_por=str(row_dict.get("creado_por") or "").strip() or None,
                    fecha_creacion=str(row_dict.get("fecha_creacion") or "").strip() or None,
                    notas=str(row_dict.get("notas") or "").strip() or None,
                    is_active=_parse_active(row_dict.get("is_active", True)),
                )
                db.add(new_seller)
                try:
                    await db.commit()
                    await db.refresh(new_seller)
                    creados += 1
                except IntegrityError:
                    await db.rollback()
                    errors.append(ImportError(fila=row_idx, seller_id=sid, motivo="id_ecommerce o seller_id ya existe"))

        except Exception as e:
            await db.rollback()
            errors.append(ImportError(fila=row_idx, seller_id=sid, motivo=str(e)))

    total = len(rows) - 1
    return SellerImportUpdateResult(
        total=total,
        actualizados=actualizados,
        creados=creados,
        errores=len(errors),
        detalle_errores=errors,
    )
