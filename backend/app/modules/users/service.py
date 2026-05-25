import io
import uuid
from datetime import datetime, timezone

import openpyxl
from fastapi import HTTPException, UploadFile, status
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.security import hash_password, verify_password
from app.modules.auth.models import User, UserRole

from .schemas import ChangePassword, SelfChangePassword, UserCreate, UserImportError, UserImportUpdateResult, UserUpdate


async def get_all_users(db: AsyncSession) -> list[User]:
    result = await db.execute(select(User).order_by(User.created_at))
    return list(result.scalars().all())


async def get_user_by_id(user_id: uuid.UUID, db: AsyncSession) -> User:
    result = await db.execute(select(User).where(User.id == user_id))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    return user


async def create_user(data: UserCreate, db: AsyncSession) -> User:
    user = User(
        username=data.username,
        email=data.email,
        full_name=data.full_name,
        hashed_password=hash_password(data.password),
        role=data.role,
        is_active=True,
    )
    db.add(user)
    try:
        await db.commit()
        await db.refresh(user)
    except IntegrityError:
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Username or email already exists",
        )
    return user


async def update_user(user_id: uuid.UUID, data: UserUpdate, db: AsyncSession) -> User:
    user = await get_user_by_id(user_id, db)
    for field, value in data.model_dump(exclude_none=True).items():
        setattr(user, field, value)
    try:
        await db.commit()
        await db.refresh(user)
    except IntegrityError:
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Email already in use",
        )
    return user


async def deactivate_user(
    user_id: uuid.UUID, current_user_id: uuid.UUID, db: AsyncSession
) -> User:
    if user_id == current_user_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Cannot deactivate your own account",
        )
    user = await get_user_by_id(user_id, db)
    user.is_active = False
    await db.commit()
    await db.refresh(user)
    return user


async def reset_password(
    user_id: uuid.UUID, data: ChangePassword, db: AsyncSession
) -> User:
    user = await get_user_by_id(user_id, db)
    user.hashed_password = hash_password(data.new_password)
    await db.commit()
    await db.refresh(user)
    return user


async def change_own_password(
    user: User, data: SelfChangePassword, db: AsyncSession
) -> None:
    if data.new_password != data.confirm_password:
        raise HTTPException(status_code=400, detail="Las contraseñas no coinciden")
    if not verify_password(data.current_password, user.hashed_password):
        raise HTTPException(status_code=400, detail="Contraseña actual incorrecta")
    user.hashed_password = hash_password(data.new_password)
    await db.commit()


_USER_EXPORT_COLS = ["username", "email", "full_name", "role", "is_active", "created_at", "last_login"]

_VALID_ROLES = {r.value for r in UserRole}


async def export_users_xlsx(db: AsyncSession) -> bytes:
    result = await db.execute(select(User).order_by(User.created_at))
    users = list(result.scalars().all())
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usuarios"
    ws.append(_USER_EXPORT_COLS)
    for u in users:
        ws.append([
            u.username,
            u.email,
            u.full_name,
            u.role.value if u.role else "",
            "activo" if u.is_active else "inactivo",
            u.created_at.strftime("%Y-%m-%d %H:%M") if u.created_at else "",
            u.last_login.strftime("%Y-%m-%d %H:%M") if u.last_login else "",
        ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def import_update_users(
    file: UploadFile, db: AsyncSession
) -> UserImportUpdateResult:
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return UserImportUpdateResult(total=0, actualizados=0, creados=0, errores=0, detalle_errores=[])

    raw_headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]

    actualizados = creados = 0
    errors: list[UserImportError] = []

    for row_idx, raw_row in enumerate(rows[1:], start=2):
        row_dict = {raw_headers[i]: raw_row[i] for i in range(min(len(raw_headers), len(raw_row)))}
        username = str(row_dict.get("username") or "").strip()
        if not username:
            continue

        email = str(row_dict.get("email") or "").strip()
        full_name = str(row_dict.get("full_name") or "").strip() or None
        role_raw = str(row_dict.get("role") or "").strip().lower()
        is_active_raw = str(row_dict.get("is_active") or "").strip().lower()

        role = role_raw if role_raw in _VALID_ROLES else None
        is_active = is_active_raw not in {"inactivo", "false", "0", "no"}

        try:
            result = await db.execute(select(User).where(User.username == username))
            existing = result.scalar_one_or_none()

            if existing:
                if email:
                    existing.email = email
                if full_name is not None:
                    existing.full_name = full_name
                if role:
                    existing.role = UserRole(role)
                existing.is_active = is_active
                try:
                    await db.commit()
                    await db.refresh(existing)
                    actualizados += 1
                except IntegrityError:
                    await db.rollback()
                    errors.append(UserImportError(fila=row_idx, username=username, motivo="Email ya existe"))
            else:
                if not email:
                    errors.append(UserImportError(fila=row_idx, username=username, motivo="Email requerido para crear usuario"))
                    continue
                if not role:
                    errors.append(UserImportError(fila=row_idx, username=username, motivo=f"Rol inválido: {role_raw!r}"))
                    continue
                new_user = User(
                    username=username,
                    email=email,
                    full_name=full_name,
                    hashed_password=hash_password("Provincia.2026"),
                    role=UserRole(role),
                    is_active=is_active,
                )
                db.add(new_user)
                try:
                    await db.commit()
                    await db.refresh(new_user)
                    creados += 1
                except IntegrityError:
                    await db.rollback()
                    errors.append(UserImportError(fila=row_idx, username=username, motivo="Username o email ya existe"))

        except Exception as e:
            await db.rollback()
            errors.append(UserImportError(fila=row_idx, username=username, motivo=str(e)))

    total = len(rows) - 1
    return UserImportUpdateResult(
        total=total,
        actualizados=actualizados,
        creados=creados,
        errores=len(errors),
        detalle_errores=errors,
    )
