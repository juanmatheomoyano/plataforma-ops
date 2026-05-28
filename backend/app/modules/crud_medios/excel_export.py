"""
Excel export para CRUD Medios — fiel al formato del script vtex_payment_extractor_3.5.
Sheets: RESUMEN, DASHBOARD_VENDEDORES, PAGOS_CONSOLIDADO, ERRORES
"""
import io
from datetime import datetime

import openpyxl
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .schemas import SellerDashboard
from .service import CUOTA_CONFIGS

# ── Estilos ───────────────────────────────────────────────────────────────────

_BORDE = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

C = {
    "azul":       PatternFill("solid", start_color="1F4E79"),
    "rojo":       PatternFill("solid", start_color="C00000"),
    "verde_h":    PatternFill("solid", start_color="375623"),
    "naranja":    PatternFill("solid", start_color="BF5700"),
    "fila_par":   PatternFill("solid", start_color="D6E4F0"),
    "fila_error": PatternFill("solid", start_color="FFE0E0"),
    "h_blanco":   Font(bold=True, color="FFFFFF", name="Arial", size=10),
    "celda":      Font(name="Arial", size=9),
    "centro":     Alignment(horizontal="center", vertical="center"),
    "izquierda":  Alignment(horizontal="left", vertical="center"),
}

VALIDATION_FILLS = {
    "Ok (vigente)":    PatternFill("solid", start_color="E2EFDA"),
    "Ok (programado)": PatternFill("solid", start_color="DDEBF7"),
    "Ok (inactiva)":   PatternFill("solid", start_color="F2F2F2"),
    "A corregir":      PatternFill("solid", start_color="FFE0E0"),
    "No configurado":  PatternFill("solid", start_color="FFFFFF"),
}
VALIDATION_FONTS = {
    "Ok (vigente)":    Font(name="Arial", size=9, bold=True, color="375623"),
    "Ok (programado)": Font(name="Arial", size=9, bold=True, color="1F4E79"),
    "Ok (inactiva)":   Font(name="Arial", size=9, color="888888"),
    "A corregir":      Font(name="Arial", size=9, bold=True, color="C00000"),
    "No configurado":  Font(name="Arial", size=9, color="AAAAAA"),
}

VALIDATION_COLS = {cfg[0] for cfg in CUOTA_CONFIGS}

COL_WIDTHS = {
    "vendedor": 24, "cuenta": 24, "id_regla": 38, "nombre_regla": 35,
    "id_sistema_pago": 18, "sistema_pago": 18, "nivel_tarjeta": 18,
    "cobrand": 14, "emisor": 38, "conector": 18, "conector_completo": 58,
    "id_afiliacion": 38, "habilitada": 12, "es_default": 12,
    "vigente_hoy": 18, "max_cuotas": 14, "cuotas_disponibles": 30,
    "tiene_interes": 14, "valor_minimo_cuota": 20, "fecha_inicio": 22,
    "fecha_fin": 22, "es_self_autorizado": 20, "requiere_autenticacion": 22,
    "servicio_cuotas": 18, "interes_externo": 16, "valor_minimo": 16,
    "canales_venta": 16, "pais": 8,
    "Seller": 26, "Seller ID": 26,
    "Tarjetas totales": 18, "Tarjetas activas": 18, "Tarjetas inactivas": 18,
    "Vigentes hoy": 16, "Firmas": 32, "Max cuotas activas": 20,
    "Conectores": 28, "Emisores": 44,
    "Tarjetas en 1 pago": 22, "Tarjetas en 6 cuotas": 22,
    "Tarjetas en 9 cuotas": 22, "Tarjetas en 12 cuotas": 22,
    "Tarjetas en 18 cuotas": 22, "Tarjetas en 24 cuotas": 22,
    "Motivo": 80,
    # Evento
    "Estado evento": 22, "Motivos evento": 80, "Reglas evento": 16,
}

PAGOS_COLS = [
    "vendedor", "cuenta", "id_regla", "nombre_regla",
    "id_sistema_pago", "sistema_pago", "nivel_tarjeta",
    "cobrand", "emisor", "conector", "conector_completo",
    "id_afiliacion", "habilitada", "es_default",
    "vigente_hoy", "max_cuotas", "cuotas_disponibles",
    "tiene_interes", "valor_minimo_cuota", "fecha_inicio",
    "fecha_fin", "es_self_autorizado", "requiere_autenticacion",
    "servicio_cuotas", "interes_externo", "valor_minimo",
    "canales_venta", "pais",
]

DASHBOARD_COLS = (
    ["Seller", "Seller ID",
     "Tarjetas totales", "Tarjetas activas", "Tarjetas inactivas",
     "Vigentes hoy", "Firmas", "Max cuotas activas", "Conectores", "Emisores"]
    + [cfg[0] for cfg in CUOTA_CONFIGS]
    + ["Motivo"]
)

ERRORES_COLS = ["vendedor", "cuenta", "error", "fecha"]


# ── Helpers ───────────────────────────────────────────────────────────────────

def _sc(ws, ref, value, bold=False, size=10, color="000000",
        bg=None, align="left", italic=False):
    cell = ws[ref]
    cell.value = value
    cell.font = Font(bold=bold, size=size, color=color, name="Arial", italic=italic)
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = _BORDE
    return cell


def _write_header_row(ws, columns: list[str], header_fill):
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22
    for col_idx, header in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = C["h_blanco"]
        cell.alignment = C["centro"]
        cell.border = _BORDE
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(
            header, max(len(header) + 4, 14)
        )


def _write_data_rows(ws, columns: list[str], rows: list[dict]):
    for r_idx, row_dict in enumerate(rows, 2):
        base_fill = C["fila_par"] if r_idx % 2 == 0 else None
        for c_idx, col_name in enumerate(columns, 1):
            val = row_dict.get(col_name, "")
            val = "" if val is None else str(val)
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = _BORDE
            if col_name in VALIDATION_COLS:
                cell.fill = VALIDATION_FILLS.get(val, PatternFill())
                cell.font = VALIDATION_FONTS.get(val, C["celda"])
                cell.alignment = C["centro"]
            else:
                cell.font = C["celda"]
                cell.alignment = C["izquierda"]
                if base_fill:
                    cell.fill = base_fill
    if len(rows) > 0:
        ws.auto_filter.ref = ws.dimensions


# ── Sheet writers ─────────────────────────────────────────────────────────────

def _write_pagos_sheet(wb, all_rows: list[dict]):
    ws = wb.create_sheet("PAGOS_CONSOLIDADO")
    _write_header_row(ws, PAGOS_COLS, C["azul"])
    _write_data_rows(ws, PAGOS_COLS, all_rows)


def _write_dashboard_sheet(wb, dashboards: list[SellerDashboard]):
    ws = wb.create_sheet("DASHBOARD_VENDEDORES")
    _write_header_row(ws, DASHBOARD_COLS, C["naranja"])

    rows: list[dict] = []
    for d in sorted(dashboards, key=lambda x: x.seller_name):
        row: dict = {
            "Seller":             d.seller_name,
            "Seller ID":          d.seller_id,
            "Tarjetas totales":   d.totales,
            "Tarjetas activas":   d.activas,
            "Tarjetas inactivas": d.inactivas,
            "Vigentes hoy":       d.vigentes_hoy,
            "Firmas":             ", ".join(d.firmas),
            "Max cuotas activas": d.max_cuotas_activas,
            "Conectores":         ", ".join(d.conectores),
            "Emisores":           ", ".join(d.emisores),
            "Motivo":             d.motivos_all,
        }
        for col_name, _, _ in CUOTA_CONFIGS:
            grupo = d.grupos.get(col_name)
            row[col_name] = grupo.estado if grupo else "No configurado"
        rows.append(row)

    _write_data_rows(ws, DASHBOARD_COLS, rows)


def _write_errores_sheet(wb, error_rows: list[dict]):
    ws = wb.create_sheet("ERRORES")
    _write_header_row(ws, ERRORES_COLS, C["rojo"])
    _write_data_rows(ws, ERRORES_COLS, error_rows)


def _write_resumen_sheet(wb, all_rows: list[dict], error_rows: list[dict],
                          elapsed: float, zero_rules: int):
    ws = wb.create_sheet("RESUMEN", 0)
    ws.sheet_view.showGridLines = False

    ok_sellers = len({r["vendedor"] for r in all_rows}) if all_rows else 0
    err_sellers = len(error_rows)
    total_rows = len(all_rows)

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 20

    ws.merge_cells("B2:C2")
    _sc(ws, "B2", "VTEX Payment Extractor — Reporte de Ejecución",
        bold=True, size=14, color="1F4E79", align="left")
    ws.row_dimensions[2].height = 26

    ws.merge_cells("B3:C3")
    _sc(ws, "B3", f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        size=9, color="888888", italic=True, align="left")
    ws.row_dimensions[3].height = 14
    ws.row_dimensions[4].height = 8

    ws.merge_cells("B5:C5")
    _sc(ws, "B5", "Resumen de Ejecución", bold=True, size=11,
        color="FFFFFF", bg="2E75B6", align="center")
    ws.row_dimensions[5].height = 22

    kpis = [
        ("Sellers procesados correctamente", ok_sellers,  "375623" if ok_sellers else None),
        ("Sellers procesados con error",     err_sellers, "C00000" if err_sellers else None),
        ("Sellers sin reglas configuradas",  zero_rules,  None),
        ("Total reglas extraídas",           total_rows,  None),
        ("Tiempo de ejecución",              f"{round(elapsed, 1)}s", None),
    ]
    for i, (label, value, val_color) in enumerate(kpis, 6):
        alt = "D6E4F0" if i % 2 == 0 else "FFFFFF"
        _sc(ws, f"B{i}", label, size=10, bg=alt)
        _sc(ws, f"C{i}", value, bold=True, size=10,
            color=val_color or "000000", bg=alt, align="center")
        ws.row_dimensions[i].height = 18

    # Pie chart
    AUX = 20
    ws[f"H{AUX}"].value = "Estado"
    ws[f"I{AUX}"].value = "Cantidad"
    ws[f"H{AUX+1}"].value = "Correctamente"
    ws[f"I{AUX+1}"].value = ok_sellers
    ws[f"H{AUX+2}"].value = "Con error"
    ws[f"I{AUX+2}"].value = err_sellers

    total = ok_sellers + err_sellers
    if total > 0:
        pie = PieChart()
        pie.title = "Sellers procesados"
        pie.style = 10
        pie.width = 13
        pie.height = 9
        pie_data = Reference(ws, min_col=9, min_row=AUX + 1, max_row=AUX + 2)
        pie_labels = Reference(ws, min_col=8, min_row=AUX + 1, max_row=AUX + 2)
        pie.add_data(pie_data)
        pie.set_categories(pie_labels)
        dp_ok = DataPoint(idx=0)
        dp_ok.graphicalProperties.solidFill = "375623"
        dp_err = DataPoint(idx=1)
        dp_err.graphicalProperties.solidFill = "C00000"
        pie.series[0].dPt.extend([dp_ok, dp_err])
        ws.add_chart(pie, "E5")


# ── Main export ───────────────────────────────────────────────────────────────

def build_excel(
    all_rows: list[dict],
    dashboards: list[SellerDashboard],
    error_rows: list[dict],
    elapsed: float,
    zero_rules: int = 0,
) -> bytes:
    """
    Genera el Excel en memoria y retorna los bytes.
    Sheets en orden: RESUMEN, DASHBOARD_VENDEDORES, PAGOS_CONSOLIDADO, ERRORES
    """
    wb = openpyxl.Workbook()
    del wb[wb.sheetnames[0]]

    if all_rows:
        _write_pagos_sheet(wb, all_rows)

    if dashboards:
        _write_dashboard_sheet(wb, dashboards)

    if error_rows:
        _write_errores_sheet(wb, error_rows)

    _write_resumen_sheet(wb, all_rows, error_rows, elapsed, zero_rules)

    # Reorder sheets
    order = ["RESUMEN", "DASHBOARD_VENDEDORES", "PAGOS_CONSOLIDADO", "ERRORES"]
    existing = [s for s in order if s in wb.sheetnames]
    for target_idx, name in enumerate(existing):
        current_idx = wb.sheetnames.index(name)
        wb.move_sheet(name, offset=target_idx - current_idx)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_excel_evento(
    results_by_seller: list[dict],
    evento_nombre: str,
    elapsed: float,
) -> bytes:
    """
    Excel de validación de evento: una hoja con estado por seller.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VALIDACION_EVENTO"

    cols = ["Seller", "Seller ID", "Estado evento", "Reglas evento", "Motivos evento"]
    _write_header_row(ws, cols, C["azul"])

    rows: list[dict] = []
    for r in results_by_seller:
        rows.append({
            "Seller":           r.get("seller_name", ""),
            "Seller ID":        r.get("seller_id", ""),
            "Estado evento":    r.get("estado_general", ""),
            "Reglas evento":    str(r.get("total_rules_evento", 0)),
            "Motivos evento":   " | ".join(r.get("motivos", [])),
        })

    _write_data_rows(ws, cols, rows)

    # Summary sheet
    ws_sum = wb.create_sheet("RESUMEN")
    ws_sum.sheet_view.showGridLines = False
    ws_sum.column_dimensions["A"].width = 3
    ws_sum.column_dimensions["B"].width = 40
    ws_sum.column_dimensions["C"].width = 20

    ws_sum.merge_cells("B2:C2")
    _sc(ws_sum, "B2", f"Validación de Evento: {evento_nombre}",
        bold=True, size=14, color="1F4E79", align="left")
    ws_sum.row_dimensions[2].height = 26
    ws_sum.merge_cells("B3:C3")
    _sc(ws_sum, "B3", f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        size=9, color="888888", italic=True, align="left")

    ok = sum(1 for r in results_by_seller if r.get("estado_general") == "Ok")
    a_corregir = sum(1 for r in results_by_seller if r.get("estado_general") == "A corregir")
    no_config = sum(1 for r in results_by_seller if r.get("estado_general") == "No configurado")

    kpis = [
        ("Sellers OK",             ok,         "375623"),
        ("Sellers a corregir",     a_corregir, "C00000" if a_corregir else None),
        ("Sellers no configurado", no_config,  None),
        ("Tiempo de ejecución",    f"{round(elapsed, 1)}s", None),
    ]
    for i, (label, value, val_color) in enumerate(kpis, 5):
        alt = "D6E4F0" if i % 2 == 0 else "FFFFFF"
        _sc(ws_sum, f"B{i}", label, size=10, bg=alt)
        _sc(ws_sum, f"C{i}", value, bold=True, size=10,
            color=val_color or "000000", bg=alt, align="center")
        ws_sum.row_dimensions[i].height = 18

    # Reorder
    wb.move_sheet("RESUMEN", offset=-wb.sheetnames.index("RESUMEN"))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
